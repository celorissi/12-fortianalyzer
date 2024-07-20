import paramiko
import time
import re
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Informações de conexão SSH
fortianalyzer_ip = '192.168.15.2'
hostname = 'fortianalyzer'
username = 'admin'
password = 'Admin@123'

# Nome do arquivo de saída
output_file = '/home/marcelo/automation/fortianalyzer/output/ssh_session.txt'
excel_file = '/home/marcelo/automation/fortianalyzer/result/results.xlsx'

# Criando o diretório caso ele não exista
os.makedirs(os.path.dirname(output_file), exist_ok=True)

# Criando uma instância do cliente SSH
ssh_client = paramiko.SSHClient()

# Ignorando o aviso de chave host desconhecida
ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

# Carregar ou criar a planilha Excel
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(["Hostname", "IP", "Data", "Resultado"])

try:
    # Conectando ao FortiAnalyzer
    ssh_client.connect(fortianalyzer_ip, username=username, password=password, timeout=5)

    # Abrindo um canal SSH
    ssh_shell = ssh_client.invoke_shell()

    # Esperando um momento para que a shell esteja pronta
    time.sleep(2)

    # Limpar buffer inicial
    output = ssh_shell.recv(65535).decode('utf-8')

    # Escrevendo o início da conexão no arquivo de saída
    with open(output_file, 'w') as f:
        f.write(output)

    # Enviando o comando "diagnose dvm task repair"
    ssh_shell.send('diagnose dvm task repair\r')
    time.sleep(2)

    # Esperando a solicitação de confirmação "(y/n)"
    while True:
        if ssh_shell.recv_ready():
            output += ssh_shell.recv(65535).decode('utf-8')
            with open(output_file, 'a') as f:
                f.write(output)
            if re.search(r'Do you want to continue\? \(y/n\)', output):
                break
        else:
            time.sleep(1)

    # Imprimindo a solicitação de confirmação
    print(output)

    # Enviando a resposta "y" para continuar
    ssh_shell.send('y\r\n')
    time.sleep(2)

    # Lendo a saída após enviar "y"
    output = ''
    while True:
        if ssh_shell.recv_ready():
            data = ssh_shell.recv(65535)
            output += data.decode('utf-8')
            with open(output_file, 'a') as f:
                f.write(data.decode('utf-8'))
            # Verifique o prompt de acordo com o seu sistema
            if re.search(r'Rebooting...', output):
                break
        else:
            time.sleep(1)

    # Imprimindo a saída final após o comando
    print(output)
    
    # Adicionando resultado à planilha
    result = "Sucesso"
    ws.append([hostname, fortianalyzer_ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result])

except paramiko.AuthenticationException:
    print("Authentication failed. Please check your credentials.")
    result = "Falha - Autenticação"
    ws.append([hostname, fortianalyzer_ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result])
except paramiko.SSHException as e:
    print(f"SSH error occurred: {e}")
    result = f"Falha - SSH error: {e}"
    ws.append([hostname, fortianalyzer_ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result])
except Exception as e:
    print(f"An error occurred: {e}")
    result = f"Falha - {e}"
    ws.append([hostname, fortianalyzer_ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result])

finally:
    # Salvando a planilha Excel
    wb.save(excel_file)
    
    # Fechando o shell SSH
    if ssh_shell:
        ssh_shell.close()
    
    # Fechando a conexão SSH
    if ssh_client:
        ssh_client.close()
