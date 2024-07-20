import paramiko
import time
import re
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Nome do arquivo de entrada com a lista de FortiAnalyzers
input_excel_file = '/home/marcelo/automation/fortianalyzer/fortianalyzers_list.xlsx'
excel_file = '/home/marcelo/automation/fortianalyzer/result/results.xlsx'
output_file_template = '/home/marcelo/automation/fortianalyzer/output/ssh_session_{hostname}.txt'

# Criando o diretório caso ele não exista
os.makedirs(os.path.dirname(excel_file), exist_ok=True)

# Criando ou carregando a planilha Excel de resultados
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(["Hostname", "IP", "Data", "Resultado"])

# Função para ler a lista de FortiAnalyzers do arquivo Excel
def load_fortianalyzers_from_excel(file_path):
    fortianalyzers = []
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        ip, hostname, username, password = row
        fortianalyzers.append({
            'ip': ip,
            'hostname': hostname,
            'username': username,
            'password': password
        })
    return fortianalyzers

# Função para executar o comando em um FortiAnalyzer
def run_diagnose(fortianalyzer):
    ip = fortianalyzer['ip']
    hostname = fortianalyzer['hostname']
    username = fortianalyzer['username']
    password = fortianalyzer['password']
    output_file = output_file_template.format(hostname=hostname)

    ssh_client = None
    ssh_shell = None

    try:
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_client.connect(ip, username=username, password=password, timeout=5)

        ssh_shell = ssh_client.invoke_shell()
        time.sleep(2)
        output = ssh_shell.recv(65535).decode('utf-8')

        with open(output_file, 'w') as f:
            f.write(output)

        ssh_shell.send('diagnose dvm task repair\r')
        time.sleep(2)

        while True:
            if ssh_shell.recv_ready():
                output += ssh_shell.recv(65535).decode('utf-8')
                with open(output_file, 'a') as f:
                    f.write(output)
                if re.search(r'Do you want to continue\? \(y/n\)', output):
                    break
            else:
                time.sleep(1)

        print(f"[{hostname}] Output before sending 'y': {output}")
        ssh_shell.send('y\r\n')
        time.sleep(2)

        output = ''
        while True:
            if ssh_shell.recv_ready():
                data = ssh_shell.recv(65535)
                output += data.decode('utf-8')
                with open(output_file, 'a') as f:
                    f.write(data.decode('utf-8'))
                if re.search(r'Rebooting...', output):
                    break
            else:
                time.sleep(1)

        print(f"[{hostname}] Final output: {output}")
        result = "Sucesso"
        ws.append([hostname, ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result])

    except paramiko.AuthenticationException:
        print(f"Authentication failed for {hostname}.")
        result = "Falha - Autenticação"
        ws.append([hostname, ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result])
    except paramiko.SSHException as e:
        print(f"SSH error occurred for {hostname}: {e}")
        result = f"Falha - SSH error: {e}"
        ws.append([hostname, ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result])
    except Exception as e:
        print(f"An error occurred for {hostname}: {e}")
        result = f"Falha - {e}"
        ws.append([hostname, ip, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result])
    finally:
        # Fechando o shell SSH e a conexão SSH se foram inicializados
        if ssh_shell:
            try:
                ssh_shell.close()
            except Exception as e:
                print(f"Failed to close SSH shell for {hostname}: {e}")
        if ssh_client:
            try:
                ssh_client.close()
            except Exception as e:
                print(f"Failed to close SSH client for {hostname}: {e}")

# Carregar a lista de FortiAnalyzers do arquivo Excel
fortianalyzers = load_fortianalyzers_from_excel(input_excel_file)

# Executar o diagnóstico para cada FortiAnalyzer
for fortianalyzer in fortianalyzers:
    print(f"Processing {fortianalyzer['hostname']}...")
    run_diagnose(fortianalyzer)
    print(f"Finished processing {fortianalyzer['hostname']}.")

# Salvando a planilha Excel com os resultados
wb.save(excel_file)
print("Results saved to Excel.")
