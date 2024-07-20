from netmiko import ConnectHandler
import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

hora_inicio = datetime.datetime.now()

##### Variáveis #####

usuario = "admin"
senha = "Admin@123"
hora = datetime.datetime.now()
planilha_origem = '/home/marcelo/automation/fortianalyzer/FG_Lojas_2021_fev.xlsx'
planilha_destino = '/home/marcelo/automation/fortianalyzer/resultado_' + hora.strftime("%Y.%m.%d_%H.%M.%S") + '.xlsx'

comandos_lista = [
    'diagnose dvm '
]

comando_validacao = "get sys interface"

##### planilha de origem #####

wb_origem = load_workbook(filename = planilha_origem)
ws_origem = wb_origem['faz']

##### planilha de destino #####

wb_destino = Workbook()
ws_destino = wb_destino.active
ws_destino.title = "Resultados"
ws_destino.cell(column=1, row=1, value="Hostname")
ws_destino.cell(column=2, row=1, value="IP")
ws_destino.cell(column=3, row=1, value="Data")
ws_destino.cell(column=4, row=1, value="Resultado")
wb_destino.save(filename = planilha_destino)

##### Tempo decorrido para carregamentos #####

hora_fim = datetime.datetime.now()
print('\n\n\nTempo decorrido para carregamentos:')
print(hora_fim - hora_inicio)

###### Aplicação das configurações nos equipamentos #####

hora_inicio = datetime.datetime.now()

#   devices é a linha da planilha de lojas (origem), um switch por linha
for devices in range (2,3): #tentar melhorar para que o range verifique o tamanho da planilha
    Device_Dict = {'device_type': 'fortinet','ip': ws_origem['C' + str(devices)].value,'username': usuario,'password': senha, 'session_log': '/home/marcelo/automation/fortianalyzer/output/output_' + ws_origem['C' + str(devices)].value + '.txt'}
    try:
        hora = datetime.datetime.now()
        net_connect = ConnectHandler(**Device_Dict)
        hostname = net_connect.find_prompt()
        hostname = str(hostname[:-1])
        print("\nConectado no firewall " + hostname)
        net_connect.send_config_set(comandos_lista)
        showvalidacao = net_connect.send_command_timing(comando_validacao, expect_string=r"(#)")
        net_connect.disconnect()
        ws_destino.cell(column=1, row=devices, value=hostname) # hostname
        ws_destino.cell(column=2, row=devices, value=ws_origem['C' + str(devices)].value) # IP
        ws_destino.cell(column=3, row=devices, value=hora) # hora
        ws_destino.cell(column=4, row=devices, value='Sucesso') # resultado
        ws_destino.cell(column=5, row=devices, value=showvalidacao) # log 
        wb_destino.save(filename = planilha_destino)
    except Exception as e:
        print('Um erro aconteceu no firewall ' + ws_origem['A' + str(devices)].value)
        print(e)
        ws_destino.cell(column=1, row=devices, value=ws_origem['A' + str(devices)].value) # hostname
        ws_destino.cell(column=2, row=devices, value=ws_origem['C' + str(devices)].value) # IP
        ws_destino.cell(column=3, row=devices, value=hora) # hora
        ws_destino.cell(column=4, row=devices, value=str(e)) # resultado
        ws_destino.cell(column=5, row=devices, value='falhou') # resultado
        wb_destino.save(filename = planilha_destino)

hora_fim = datetime.datetime.now()
print('\n\n\nTempo decorrido nos equipamentos:')
print(hora_fim - hora_inicio)
print('\n\n\nFim do script')
